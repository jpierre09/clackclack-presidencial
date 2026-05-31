"""
Corre OCR local sobre uno o varios PDFs de E-14 presidencial y exporta
los resultados a Excel con columnas:

  PDF | Pagina | Region | Tipo | Numero | Nombre | OCR_valor | OCR_raw | OCR_conf | Real_valor | Diferencia | OK

Uso:
    python tools/ocr_to_excel.py
    python tools/ocr_to_excel.py --pdf "E14 TEST Presidencial/E14_XXX_X_60_010_000_00_000_X_XXX.pdf"
    python tools/ocr_to_excel.py --all          # procesa todos los PDFs de test
    python tools/ocr_to_excel.py --out resultados_ocr.xlsx

La columna "Real_valor" queda vacía — la llenas manualmente en Excel.
La columna "Diferencia" y "OK" se calculan con fórmulas Excel automáticamente.
"""
from __future__ import annotations

import argparse
import re
import sys
import time
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import fitz
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Font, PatternFill, Side, Border)
from openpyxl.utils import get_column_letter

from backend.services.local_ocr import (
    crop_region, crop_vote_column, detect_six_signatures, detect_recount,
    load_template_regions,
    process_blancos_nulos, process_nivelacion, recognize_number,
    render_page, suppress_horizontal_lines,
)

# ── Output directory ──────────────────────────────────────────────────────────
EXPORTS_DIR = ROOT / "exports"


def _find_test_pdfs() -> list[Path]:
    test_dir = ROOT / "E14 TEST Presidencial"
    if test_dir.exists():
        return sorted(test_dir.glob("*.pdf"))
    return []


# ── OCR one PDF, returning per-region rows ────────────────────────────────────
def ocr_pdf(pdf_path: Path) -> list[dict]:
    regions = load_template_regions()
    if not regions:
        print("  AVISO: no hay regiones en data/e14_template.json — ejecuta region_selector.py primero")
        return []

    doc = fitz.open(str(pdf_path))
    total_pages = len(doc)
    doc.close()

    pages_needed = sorted(set(r["page"] for r in regions if r["page"] <= total_pages))
    rendered: dict[int, np.ndarray] = {}
    for pn in pages_needed:
        rendered[pn] = render_page(str(pdf_path), pn, scale=2.0)

    rows: list[dict] = []

    for region in regions:
        pn = region["page"]
        if pn not in rendered:
            continue

        page_img = rendered[pn]
        crop = crop_region(page_img, region)
        tipo = region.get("tipo", "otro")
        label = region.get("label", tipo)
        numero = region.get("numero")
        nombre = region.get("nombre") or label

        base_row = {
            "pdf": pdf_path.name,
            "pagina": pn,
            "region_id": region.get("id", ""),
            "tipo": tipo,
            "numero": numero,
            "nombre": nombre,
            "ocr_valor": None,
            "ocr_raw": "",
            "ocr_conf": None,
        }

        if tipo == "candidato":
            # Usamos el crop tal como fue seleccionado por el usuario
            clean = suppress_horizontal_lines(crop)
            val, conf, raw = recognize_number(clean)
            rows.append({**base_row, "ocr_valor": val, "ocr_raw": raw, "ocr_conf": round(conf * 100)})

        elif tipo == "nivelacion":
            # La region YA es solo la cajita de digitos — OCR directo
            clean = suppress_horizontal_lines(crop)
            val, conf, raw = recognize_number(clean)
            rows.append({**base_row, "ocr_valor": val, "ocr_raw": raw, "ocr_conf": round(conf * 100)})

        elif tipo == "blancos_nulos":
            # La region YA es solo la cajita de digitos — OCR directo
            clean = suppress_horizontal_lines(crop)
            val, conf, raw = recognize_number(clean)
            rows.append({**base_row, "ocr_valor": val, "ocr_raw": raw, "ocr_conf": round(conf * 100)})

        elif tipo == "firmas":
            firmas = detect_six_signatures(crop)
            for i, present in enumerate(firmas, 1):
                rows.append({**base_row, "region_id": region.get("id","")+"_"+str(i),
                              "numero": f"FIRMA{i}", "nombre": f"Firma Jurado {i}",
                              "ocr_valor": 1 if present else 0,
                              "ocr_raw": "PRESENTE" if present else "AUSENTE",
                              "ocr_conf": None})

        elif tipo == "recuento":
            resultado = detect_recount(crop)
            rows.append({**base_row,
                         "numero": "RECUENTO", "nombre": "Hubo recuento de votos",
                         "ocr_valor": "SI" if resultado is True else ("NO" if resultado is False else "?"),
                         "ocr_raw": str(resultado), "ocr_conf": None})

    return rows


# ── Build Excel workbook ──────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
CAND_FILL     = PatternFill("solid", fgColor="0D2010")
NIV_FILL      = PatternFill("solid", fgColor="0D1540")
BLANCOS_FILL  = PatternFill("solid", fgColor="2A2000")
FIRMA_FILL    = PatternFill("solid", fgColor="1A0030")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
DATA_FONT     = Font(name="Calibri", size=10)
NUMBER_FONT   = Font(bold=True, name="Calibri", size=11)
THIN          = Side(style="thin", color="333333")
THIN_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TIPO_FILL = {
    "candidato":    CAND_FILL,
    "nivelacion":   NIV_FILL,
    "blancos_nulos": BLANCOS_FILL,
    "firmas":       FIRMA_FILL,
    "recuento":     PatternFill("solid", fgColor="2A0020"),
}


def build_excel(all_rows: list[dict], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR vs Real"

    headers = [
        "PDF", "Pag", "Tipo", "N°", "Nombre / Campo",
        "OCR_valor", "OCR_raw", "OCR_conf%",
        "Real_valor", "Diferencia", "OK",
    ]
    col_widths = [28, 5, 14, 6, 32, 11, 14, 11, 11, 12, 6]

    # Header row
    ws.row_dimensions[1].height = 20
    for col, (hdr, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(all_rows, start=2):
        tipo = row.get("tipo", "otro")
        fill = TIPO_FILL.get(tipo, PatternFill("solid", fgColor="1A1A1A"))

        values = [
            row.get("pdf", ""),
            row.get("pagina", ""),
            tipo,
            row.get("numero", ""),
            row.get("nombre", ""),
            row.get("ocr_valor"),
            row.get("ocr_raw", ""),
            row.get("ocr_conf"),
            None,  # Real_valor — user fills this
            None,  # Diferencia — formula
            None,  # OK — formula
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = DATA_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

        # Numeric columns center
        for col in (2, 4, 6, 8, 9):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Diferencia formula: Real - OCR (col I - col F)
        ocr_col = "F"
        real_col = "I"
        diff_cell = ws.cell(row=row_idx, column=10)
        ok_cell   = ws.cell(row=row_idx, column=11)

        # Only add formula when both are likely numbers (candidato / nivelacion / blancos)
        if tipo in ("candidato", "nivelacion", "blancos_nulos"):
            diff_cell.value = f"=IF(AND({real_col}{row_idx}<>\"\"," \
                              f"{ocr_col}{row_idx}<>\"\")," \
                              f"{real_col}{row_idx}-{ocr_col}{row_idx},\"\")"
            ok_cell.value  = f"=IF({ocr_col}{row_idx}={real_col}{row_idx},\"OK\",\"\")"

    # ── Second sheet: Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "Generado"
    ws2["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws2["A2"] = "Total filas"
    ws2["B2"] = len(all_rows)
    ws2["A3"] = "PDFs procesados"
    ws2["B3"] = len(set(r["pdf"] for r in all_rows))
    ws2["A4"] = "Candidatos"
    ws2["B4"] = len([r for r in all_rows if r["tipo"] == "candidato"])

    ws2["A6"] = "INSTRUCCIONES"
    ws2["A6"].font = Font(bold=True)
    instructions = [
        "1. Rellena la columna 'Real_valor' en la hoja 'OCR vs Real' con los valores correctos del acta.",
        "2. La columna 'Diferencia' se calcula automáticamente (Real - OCR).",
        "3. La columna 'OK' muestra 'OK' cuando OCR == Real.",
        "4. Ajusta las regiones en data/e14_template.json (usa tools/region_selector.py).",
        "5. Vuelve a ejecutar este script para ver si mejora.",
    ]
    for i, inst in enumerate(instructions, 7):
        ws2.cell(row=i, column=1, value=inst)
    ws2.column_dimensions["A"].width = 70

    wb.save(str(out_path))
    print(f"\nExcel guardado: {out_path}")
    print(f"  Filas: {len(all_rows)}  |  Abre el archivo y rellena la columna 'Real_valor'")


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="OCR E-14 → Excel comparativo")
    parser.add_argument("--pdf", default=None, help="PDF específico a procesar")
    parser.add_argument("--all", action="store_true", help="Procesar todos los PDFs de test")
    parser.add_argument("--out", default=None, help="Archivo de salida .xlsx")
    args = parser.parse_args()

    EXPORTS_DIR.mkdir(exist_ok=True)

    if args.pdf:
        pdf_path = Path(args.pdf)
        if not pdf_path.is_absolute():
            pdf_path = ROOT / args.pdf
        pdfs = [pdf_path]
    elif args.all:
        pdfs = _find_test_pdfs()
        if not pdfs:
            print("ERROR: No se encontraron PDFs de test.")
            sys.exit(1)
    else:
        pdfs = _find_test_pdfs()[:1]  # default: first test PDF
        if not pdfs:
            print("ERROR: No se encontraron PDFs de test en 'E14 TEST Presidencial/'")
            sys.exit(1)

    out_name = args.out or f"ocr_comparacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    out_path = EXPORTS_DIR / out_name if not Path(out_name).is_absolute() else Path(out_name)

    all_rows: list[dict] = []
    for pdf in pdfs:
        print(f"\nProcesando: {pdf.name} ...", flush=True)
        t0 = time.time()
        rows = ocr_pdf(pdf)
        elapsed = time.time() - t0
        all_rows.extend(rows)
        print(f"  {len(rows)} filas extraidas en {elapsed:.1f}s")

    if not all_rows:
        print("No se extrajeron datos.")
        sys.exit(1)

    build_excel(all_rows, out_path)


if __name__ == "__main__":
    main()
