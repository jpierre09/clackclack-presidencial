"""Test E-14 PDF processing with Claude and export results to Excel for manual validation.

Usage:
    python test_e14_to_excel.py <pdf_path> [--max-pages N]
    python test_e14_to_excel.py  (uses default test PDF)

Output: results/validation_<timestamp>.xlsx
"""
import json
import os
import sys
from datetime import datetime
from pathlib import Path

# Ensure project root is in path
sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.services.claude_ocr import process_e14_pdf, normalize_result


# --- Styling ---
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
META_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
PH_FILL = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
LOW_CONF_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INCONSISTENT_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
MANUAL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
ALERT_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _apply_header(ws, row, cols):
    for col_idx, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_cell(ws, row, col, value, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    if fill:
        cell.fill = fill
    return cell


def result_to_excel(result: dict, pdf_path: str, output_path: str):
    """Convert Claude extraction result to a validation Excel file.

    Handles both the new detailed format (with encabezado/candidatos/reconciliacion)
    and the normalized flat format.
    """
    # Normalize if needed
    if "encabezado" in result:
        norm = normalize_result(result)
        raw_result = result  # Keep raw for JSON sheet
    else:
        norm = result
        raw_result = result

    wb = Workbook()

    # --- Sheet 1: Resumen ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.column_dimensions["A"].width = 25
    ws_resumen.column_dimensions["B"].width = 45

    niv = norm.get("nivelacion", {})
    recon = norm.get("_reconciliacion", raw_result.get("reconciliacion", {}))
    meta = norm.get("_meta", {})

    info_fields = [
        ("PDF", os.path.basename(pdf_path)),
        ("Serial", norm.get("serial", "")),
        ("Departamento", f"{norm.get('departamento_cod', '')} - {norm.get('departamento_nombre', '')}"),
        ("Municipio", f"{norm.get('municipio_cod', '')} - {norm.get('municipio_nombre', '')}"),
        ("Zona", norm.get("zona", "")),
        ("Puesto", norm.get("puesto", "")),
        ("Mesa", norm.get("mesa", "")),
        ("Corporacion", norm.get("corporacion", "")),
        ("Lugar", norm.get("lugar", "")),
        ("Codigo Transmision", norm.get("codigo_transmision", "")),
        ("", ""),
        ("NIVELACION", ""),
        ("Sufragantes E-11", niv.get("total_sufragantes_e11")),
        ("Votos en Urna", niv.get("total_votos_urna")),
        ("", ""),
        ("TOTALES ESPECIALES", ""),
        ("Votos en Blanco", norm.get("votos_en_blanco")),
        ("Votos Nulos", norm.get("votos_nulos")),
        ("Votos No Marcados", norm.get("votos_no_marcados")),
        ("", ""),
        ("RECONCILIACION", ""),
        ("Suma Partidos", recon.get("suma_partidos")),
        ("Suma Especiales (blanco+nulo+nomarc)", recon.get("suma_especiales")),
        ("Suma Total Calculada", recon.get("suma_total")),
        ("Total Urna Registrado", recon.get("total_urna")),
        ("Diferencia", recon.get("diferencia")),
        ("Suma <= Urna", "SI" if recon.get("suma_menor_o_igual_urna") else "NO" if recon.get("suma_menor_o_igual_urna") is not None else recon.get("consistente", "")),
        ("Nota Reconciliacion", recon.get("nota", "")),
        ("", ""),
        ("VALIDACION POST-OCR", ""),
        ("Nivel Alerta", norm.get("_validacion", {}).get("nivel_alerta", "?")),
        ("", ""),
        ("METADATA", ""),
        ("Confianza General", norm.get("confianza_general")),
        ("Modelo", meta.get("model", "")),
        ("Paginas Enviadas", meta.get("pages_sent")),
        ("Paginas Totales PDF", meta.get("pages_total")),
        ("Tiempo API (s)", meta.get("api_time_s")),
        ("Tiempo Total (s)", meta.get("total_time_s")),
        ("Tokens Input", meta.get("input_tokens")),
        ("Tokens Output", meta.get("output_tokens")),
        ("Costo USD", meta.get("cost_total_usd")),
    ]

    for row_idx, (label, value) in enumerate(info_fields, 1):
        cell_a = ws_resumen.cell(row=row_idx, column=1, value=label)
        cell_b = ws_resumen.cell(row=row_idx, column=2, value=value)
        cell_a.font = Font(bold=True)
        if label in ("NIVELACION", "TOTALES ESPECIALES", "METADATA", "RECONCILIACION", "VALIDACION POST-OCR"):
            cell_a.fill = META_FILL
            cell_b.fill = META_FILL
        if label == "Suma <= Urna" and value == "NO":
            cell_b.fill = INCONSISTENT_FILL
        elif label == "Suma <= Urna" and value == "SI":
            cell_b.fill = OK_FILL
        if label == "Nivel Alerta":
            if value == "REQUIERE_REVISION_MANUAL":
                cell_b.fill = MANUAL_FILL
                cell_b.font = ALERT_FONT
            elif value == "ALERTA_ARITMETICA":
                cell_b.fill = ALERT_FILL
                cell_b.font = ALERT_FONT
            elif value == "OK":
                cell_b.fill = OK_FILL

    # Validation columns
    ws_resumen.column_dimensions["C"].width = 15
    ws_resumen.column_dimensions["D"].width = 20
    ws_resumen.cell(row=1, column=3, value="CORRECTO?").font = Font(bold=True, color="FF0000")
    ws_resumen.cell(row=1, column=4, value="VALOR REAL").font = Font(bold=True, color="FF0000")

    # --- Sheet 2: Partidos ---
    ws_partidos = wb.create_sheet("Partidos")
    headers = ["#", "Codigo", "Nombre Partido", "Tipo Lista",
               "Votos Lista", "Total Votos", "Suma Calc", "Consist?",
               "Confianza", "CORRECTO?", "LISTA REAL", "TOTAL REAL", "NOTAS"]
    _apply_header(ws_partidos, 1, headers)

    ws_partidos.column_dimensions["A"].width = 5
    ws_partidos.column_dimensions["B"].width = 8
    ws_partidos.column_dimensions["C"].width = 45
    ws_partidos.column_dimensions["D"].width = 20
    ws_partidos.column_dimensions["E"].width = 12
    ws_partidos.column_dimensions["F"].width = 12
    ws_partidos.column_dimensions["G"].width = 10
    ws_partidos.column_dimensions["H"].width = 10
    ws_partidos.column_dimensions["I"].width = 10
    ws_partidos.column_dimensions["J"].width = 12
    ws_partidos.column_dimensions["K"].width = 12
    ws_partidos.column_dimensions["L"].width = 12
    ws_partidos.column_dimensions["M"].width = 30

    partidos = norm.get("partidos", [])
    for idx, p in enumerate(partidos, 1):
        row = idx + 1
        nombre = p.get("nombre", "")
        confianza = p.get("confianza", 0)
        consistente = p.get("consistente")

        is_ph = "PACTO" in nombre.upper() or "HIST" in nombre.upper()
        if is_ph:
            fill = PH_FILL
        elif consistente is False:
            fill = INCONSISTENT_FILL
        elif confianza and confianza < 70:
            fill = LOW_CONF_FILL
        else:
            fill = None

        _write_cell(ws_partidos, row, 1, idx, fill)
        _write_cell(ws_partidos, row, 2, p.get("codigo", ""), fill)
        cell_name = _write_cell(ws_partidos, row, 3, nombre, fill)
        cell_name.alignment = Alignment(horizontal="left")
        _write_cell(ws_partidos, row, 4, p.get("tipo_lista", ""), fill)
        _write_cell(ws_partidos, row, 5, p.get("votos_lista"), fill)
        _write_cell(ws_partidos, row, 6, p.get("total_votos"), fill)
        _write_cell(ws_partidos, row, 7, p.get("suma_calculada"), fill)
        consist_text = "SI" if consistente else ("NO" if consistente is False else "")
        _write_cell(ws_partidos, row, 8, consist_text, fill)
        _write_cell(ws_partidos, row, 9, confianza, fill)
        for col in range(10, 14):
            _write_cell(ws_partidos, row, col, None, fill)

    # Summary row
    summary_row = len(partidos) + 3
    ws_partidos.cell(row=summary_row, column=3, value="TOTAL VOTOS PARTIDOS").font = Font(bold=True)
    total_formula = f"=SUM(F2:F{len(partidos) + 1})"
    ws_partidos.cell(row=summary_row, column=6, value=total_formula).font = Font(bold=True)

    # --- Sheet 3: Candidatos (detail per party) ---
    ws_cand = wb.create_sheet("Candidatos")
    cand_headers = ["Partido", "Cod Partido", "Cod Candidato", "Votos", "Nota"]
    _apply_header(ws_cand, 1, cand_headers)
    ws_cand.column_dimensions["A"].width = 35
    ws_cand.column_dimensions["B"].width = 12
    ws_cand.column_dimensions["C"].width = 15
    ws_cand.column_dimensions["D"].width = 10
    ws_cand.column_dimensions["E"].width = 50

    cand_row = 2
    for p in partidos:
        candidatos = p.get("candidatos", [])
        for c in candidatos:
            _write_cell(ws_cand, cand_row, 1, p.get("nombre", ""))
            ws_cand.cell(row=cand_row, column=1).alignment = Alignment(horizontal="left")
            _write_cell(ws_cand, cand_row, 2, p.get("codigo", ""))
            _write_cell(ws_cand, cand_row, 3, c.get("codigo"))
            _write_cell(ws_cand, cand_row, 4, c.get("votos"))
            nota = c.get("nota", "")
            _write_cell(ws_cand, cand_row, 5, nota)
            ws_cand.cell(row=cand_row, column=5).alignment = Alignment(horizontal="left")
            cand_row += 1

    if cand_row == 2:
        ws_cand.cell(row=2, column=1, value="(No se reportaron candidatos con votos)")

    # --- Sheet 4: Validacion Aritmetica ---
    validacion = norm.get("_validacion", {})
    ws_val = wb.create_sheet("Validacion")
    val_headers = ["Partido", "Codigo", "Nivel Alerta", "Suma Calc", "Total Reg",
                   "Diferencia", "Detalle"]
    _apply_header(ws_val, 1, val_headers)
    ws_val.column_dimensions["A"].width = 40
    ws_val.column_dimensions["B"].width = 10
    ws_val.column_dimensions["C"].width = 25
    ws_val.column_dimensions["D"].width = 12
    ws_val.column_dimensions["E"].width = 12
    ws_val.column_dimensions["F"].width = 12
    ws_val.column_dimensions["G"].width = 60

    val_row = 2
    for ap in validacion.get("alertas_partidos", []):
        nivel = ap.get("nivel", "")
        if nivel == "REQUIERE_REVISION_MANUAL":
            fill = MANUAL_FILL
            font = ALERT_FONT
        elif nivel == "ALERTA_ARITMETICA":
            fill = ALERT_FILL
            font = ALERT_FONT
        else:
            fill = OK_FILL
            font = None

        _write_cell(ws_val, val_row, 1, ap.get("nombre", ""))
        ws_val.cell(row=val_row, column=1).alignment = Alignment(horizontal="left")
        _write_cell(ws_val, val_row, 2, ap.get("codigo", ""))
        cell_nivel = _write_cell(ws_val, val_row, 3, nivel, fill)
        if font:
            cell_nivel.font = font
        _write_cell(ws_val, val_row, 4, ap.get("suma_calculada"))
        _write_cell(ws_val, val_row, 5, ap.get("total_registrado"))
        _write_cell(ws_val, val_row, 6, ap.get("diferencia"))
        _write_cell(ws_val, val_row, 7, ap.get("detalle", ""))
        ws_val.cell(row=val_row, column=7).alignment = Alignment(horizontal="left")
        val_row += 1

    # Global validation row
    ag = validacion.get("alerta_global", {})
    val_row += 1
    ws_val.cell(row=val_row, column=1, value="VALIDACION GLOBAL").font = Font(bold=True)
    ws_val.cell(row=val_row, column=1).fill = META_FILL
    for c in range(2, 8):
        ws_val.cell(row=val_row, column=c).fill = META_FILL
    val_row += 1

    g_nivel = ag.get("nivel", "")
    if g_nivel == "REQUIERE_REVISION_MANUAL":
        g_fill = MANUAL_FILL
        g_font = ALERT_FONT
    elif g_nivel == "ALERTA_ARITMETICA":
        g_fill = ALERT_FILL
        g_font = ALERT_FONT
    else:
        g_fill = OK_FILL
        g_font = None

    _write_cell(ws_val, val_row, 1, "Suma Total vs Urna")
    ws_val.cell(row=val_row, column=1).alignment = Alignment(horizontal="left")
    cell_g = _write_cell(ws_val, val_row, 3, g_nivel, g_fill)
    if g_font:
        cell_g.font = g_font
    _write_cell(ws_val, val_row, 4, ag.get("suma_total"))
    _write_cell(ws_val, val_row, 5, ag.get("total_urna"))
    _write_cell(ws_val, val_row, 6, ag.get("diferencia"))
    _write_cell(ws_val, val_row, 7, ag.get("detalle", ""))
    ws_val.cell(row=val_row, column=7).alignment = Alignment(horizontal="left")

    # Overall alert level
    val_row += 2
    nivel_general = validacion.get("nivel_alerta", "?")
    ws_val.cell(row=val_row, column=1, value="NIVEL ALERTA GENERAL:").font = Font(bold=True, size=14)
    cell_final = ws_val.cell(row=val_row, column=3, value=nivel_general)
    cell_final.font = Font(bold=True, size=14)
    if nivel_general == "REQUIERE_REVISION_MANUAL":
        cell_final.fill = MANUAL_FILL
        cell_final.font = Font(bold=True, size=14, color="FFFFFF")
    elif nivel_general == "ALERTA_ARITMETICA":
        cell_final.fill = ALERT_FILL
        cell_final.font = Font(bold=True, size=14, color="FFFFFF")
    else:
        cell_final.fill = OK_FILL

    # --- Sheet 5: JSON Raw ---
    ws_json = wb.create_sheet("JSON Raw")
    ws_json.column_dimensions["A"].width = 120
    raw_json = json.dumps(raw_result, indent=2, ensure_ascii=False)
    for line_idx, line in enumerate(raw_json.split("\n"), 1):
        ws_json.cell(row=line_idx, column=1, value=line).font = Font(name="Consolas", size=9)

    wb.save(output_path)
    return output_path


def main():
    # Default test PDF
    default_pdf = r"D:\Personal\Todo\Proyectos\Proyectos 2026\AuditorDiaDPH\E14_Downloads\Pruebas\6241727_E14_CAM_X_01_001_001_XX_05_005_X_XXX.pdf"

    pdf_path = sys.argv[1] if len(sys.argv) > 1 else default_pdf
    max_pages = 3  # Default for CAM

    model = None  # Use default from claude_ocr

    # Parse optional args
    for i, arg in enumerate(sys.argv):
        if arg == "--max-pages" and i + 1 < len(sys.argv):
            max_pages = int(sys.argv[i + 1])
        if arg == "--model" and i + 1 < len(sys.argv):
            model = sys.argv[i + 1]

    # Auto-detect corporacion from filename
    basename = os.path.basename(pdf_path).upper()
    if "SEN" in basename:
        max_pages = 10
        print(f"Detected SENADO - using max_pages={max_pages}")
    elif "CAM" in basename:
        max_pages = 3
        print(f"Detected CAMARA - using max_pages={max_pages}")

    if not os.path.exists(pdf_path):
        print(f"ERROR: PDF not found: {pdf_path}")
        sys.exit(1)

    model_display = model or "claude-sonnet-4 (default)"
    print(f"Processing: {os.path.basename(pdf_path)}")
    print(f"Settings: max_pages={max_pages}, model={model_display}")
    print("Calling Claude API...")

    kwargs = {"max_pages": max_pages}
    if model:
        kwargs["model"] = model
    raw_result = process_e14_pdf(pdf_path, **kwargs)
    norm = normalize_result(raw_result)

    # Create output directory
    results_dir = Path(__file__).parent / "results"
    results_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    corp = norm.get("corporacion", "UNK")[:3]
    mesa = norm.get("mesa", "000")
    output_name = f"validation_{corp}_{mesa}_{timestamp}.xlsx"
    output_path = results_dir / output_name

    result_to_excel(raw_result, pdf_path, str(output_path))

    # Print summary
    meta = norm.get("_meta", {})
    partidos = norm.get("partidos", [])
    recon = norm.get("_reconciliacion", raw_result.get("reconciliacion", {}))
    ph = [p for p in partidos if "PACTO" in (p.get("nombre") or "").upper()]

    print(f"\n{'='*60}")
    print(f"Corporacion: {norm.get('corporacion', '?')}")
    print(f"Mesa: {norm.get('mesa', '?')}")
    print(f"Nivelacion: E11={norm.get('nivelacion',{}).get('total_sufragantes_e11')}, Urna={norm.get('nivelacion',{}).get('total_votos_urna')}")
    print(f"Partidos: {len(partidos)}")
    print()

    for p in partidos:
        nombre = (p.get("nombre") or "")[:35]
        consist = "OK" if p.get("consistente") else "!!" if p.get("consistente") is False else "?"
        cand_count = len(p.get("candidatos", []))
        print(f"  {p.get('codigo','?'):>5} | {nombre:<35} | lista={p.get('votos_lista',0):>3} | total={p.get('total_votos',0):>3} | cand={cand_count:>2} | {consist}")

    print()
    if ph:
        print(f"PH: votos_lista={ph[0].get('votos_lista')}, total={ph[0].get('total_votos')}")

    print(f"\nReconciliacion: suma={recon.get('suma_total','?')} vs urna={recon.get('total_urna','?')} (suma <= urna: {'SI' if recon.get('suma_menor_o_igual_urna') else 'NO'})")
    if recon.get("nota"):
        print(f"  Nota: {recon['nota']}")

    # Validation alerts
    validacion = norm.get("_validacion", {})
    nivel = validacion.get("nivel_alerta", "?")
    print(f"\n--- VALIDACION POST-OCR: {nivel} ---")
    for ap in validacion.get("alertas_partidos", []):
        if ap.get("nivel") != "OK":
            print(f"  {ap['codigo']:>5} | {ap['nombre'][:30]:<30} | {ap['nivel']} | diff={ap['diferencia']} | {ap['detalle']}")
    ag = validacion.get("alerta_global", {})
    if ag.get("nivel") != "OK":
        print(f"  GLOBAL | {ag['nivel']} | {ag['detalle']}")

    print(f"\nConfianza: {norm.get('confianza_general', '?')}%")
    print(f"Tiempo: {meta.get('total_time_s', '?')}s | Costo: ${meta.get('cost_total_usd', '?')} USD")
    print(f"\nExcel: {output_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
