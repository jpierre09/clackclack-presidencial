"""Tinder-style manual validation tool API."""
from __future__ import annotations

import hashlib
import secrets
from typing import Annotated

import io
from datetime import datetime

from fastapi import APIRouter, Depends, Header, HTTPException, Response
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from backend import database as db
from backend.config import (
    BASE_DIR,
    E14_DOWNLOADS_DIR,
    PUBLIC_EXPORT_SHARE_TOKEN,
    VALIDATE_SETUP_TOKEN,
)

# Exports dir sits next to e14_downloads, so it lives on /persist on Railway
_EXPORTS_DIR = E14_DOWNLOADS_DIR.parent / "exports"
from backend.services import alert_engine
from backend.services.event_bus import event_bus

router = APIRouter(prefix="/api/validar", tags=["manual-validate"])

_ITERATIONS = 260_000

# ── Screenshot in-memory cache ────────────────────────────────────────────────
# Key: "{mun}:{zona}:{puesto}:{mesa}:{corp}" (URL params) → PNG bytes
# Checked BEFORE any DB query so cache hits cost ~0ms.
# Evicts oldest entry when full (FIFO).
_SCREENSHOT_CACHE: dict[str, bytes] = {}
_SCREENSHOT_CACHE_MAX = 500

# Dedicated thread pool for screenshot rendering — isolated from the default
# executor used by OCR workers so renders never queue behind OCR.
import concurrent.futures as _cf
_SCREENSHOT_POOL = _cf.ThreadPoolExecutor(max_workers=4, thread_name_prefix="ss")


def _safe_within(path, root) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


def _clean_zip_segment(value: str | None, fallback: str) -> str:
    import re as _re
    import unicodedata as _unicodedata

    normalized = _unicodedata.normalize("NFKD", value or "")
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = _re.sub(r"[^A-Za-z0-9._ -]+", "", ascii_value)
    cleaned = _re.sub(r"\s+", " ", cleaned).strip(" ._-")
    return cleaned or fallback


def _resolve_download_filepath(raw: str | None):
    from pathlib import Path as _Path
    import os as _os

    if not raw:
        return None

    if _os.path.isabs(raw):
        full_path = _Path(raw).resolve()
    else:
        full_path = (BASE_DIR / raw).resolve()

    allowed = [E14_DOWNLOADS_DIR.resolve(), BASE_DIR.resolve()]
    if not any(_safe_within(full_path, root) for root in allowed):
        return None
    return full_path


def _zip_stream_response(zip_path, filename: str) -> StreamingResponse:
    def _iter_file():
        with open(zip_path, "rb") as f:
            while chunk := f.read(65536):
                yield chunk

    return StreamingResponse(
        _iter_file(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _require_public_export_access(share_token: str) -> None:
    if not PUBLIC_EXPORT_SHARE_TOKEN:
        raise HTTPException(status_code=404, detail="No encontrado")
    if not share_token or not secrets.compare_digest(share_token, PUBLIC_EXPORT_SHARE_TOKEN):
        raise HTTPException(status_code=404, detail="No encontrado")


def _hash_password(password: str) -> str:
    salt = secrets.token_hex(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), _ITERATIONS)
    return f"pbkdf2:{salt}:{dk.hex()}"


def _verify_password(password: str, stored: str) -> bool:
    try:
        _, salt, dk_hex = stored.split(":", 2)
    except ValueError:
        return False
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), _ITERATIONS)
    return secrets.compare_digest(dk.hex(), dk_hex)


async def _require_auth(x_session_token: Annotated[str | None, Header()] = None) -> str:
    if not x_session_token:
        raise HTTPException(status_code=401, detail="No session token")
    username = await db.get_session(x_session_token)
    if not username:
        raise HTTPException(status_code=401, detail="Invalid or expired session")
    return username


class LoginRequest(BaseModel):
    username: str
    password: str


class CreateUserRequest(BaseModel):
    username: str
    password: str
    setup_token: str


class AdminUsersRequest(BaseModel):
    admin_token: str
    usernames: list[str] | None = None


class SubmitRequest(BaseModel):
    municipio_cod: str
    zona_cod: str
    puesto_cod: str
    mesa: int
    corporacion: str
    action: str               # "approved" | "corrected"
    corrected_ph_votes: int | None = None


class AdminCropRequest(BaseModel):
    municipio_cod: str
    zona_cod: str
    puesto_cod: str
    mesa: int
    corporacion: str
    x0: float
    y0: float
    x1: float
    y1: float
    corrected_ph_votes: int | None = None
    admin_token: str


class NoveltyRequest(BaseModel):
    municipio_cod: str
    zona_cod: str
    puesto_cod: str
    mesa: int
    corporacion: str
    note: str


# ── Auth ──────────────────────────────────────────────────────────────────────

@router.post("/auth/login")
async def login(req: LoginRequest):
    user = await db.get_user(req.username)
    if not user or not user["is_active"]:
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    if not _verify_password(req.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    token = secrets.token_urlsafe(32)
    await db.create_session(token, req.username)
    return {"token": token, "username": req.username}


@router.post("/auth/logout")
async def logout(x_session_token: Annotated[str | None, Header()] = None):
    if x_session_token:
        username = await db.get_session(x_session_token)
        if username:
            await db.release_claim(username)
        await db.delete_session(x_session_token)
    return {"status": "ok"}


@router.get("/auth/me")
async def me(username: str = Depends(_require_auth)):
    return {"username": username}


# ── Admin ─────────────────────────────────────────────────────────────────────

@router.post("/admin/crop-override")
async def admin_crop_override(req: AdminCropRequest):
    """Set a manual crop override for a PDF (admin only)."""
    if not VALIDATE_SETUP_TOKEN:
        raise HTTPException(status_code=503, detail="VALIDATE_SETUP_TOKEN not configured")
    if not secrets.compare_digest(req.admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")

    corp = req.corporacion.upper()
    await db.save_crop_override(
        req.municipio_cod, req.zona_cod, req.puesto_cod,
        req.mesa, corp, "admin",
        req.x0, req.y0, req.x1, req.y1,
    )

    if req.corrected_ph_votes is not None:
        await db.submit_validation({
            "municipio_cod": req.municipio_cod,
            "zona_cod": req.zona_cod,
            "puesto_cod": req.puesto_cod,
            "mesa": req.mesa,
            "corporacion": corp,
            "validated_by": "admin",
            "action": "corrected",
            "corrected_ph_votes": req.corrected_ph_votes,
        })
        await alert_engine.evaluate_mesa(
            req.municipio_cod, req.zona_cod, req.puesto_cod, req.mesa
        )

    return {"status": "ok"}


@router.post("/admin/create-user")
async def create_user(req: CreateUserRequest):
    if not VALIDATE_SETUP_TOKEN:
        raise HTTPException(status_code=503, detail="VALIDATE_SETUP_TOKEN not configured")
    if not secrets.compare_digest(req.setup_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Invalid setup token")
    if len(req.username) < 2 or len(req.password) < 6:
        raise HTTPException(status_code=400, detail="username >= 2 chars, password >= 6 chars")
    ok = await db.create_user(req.username, _hash_password(req.password))
    if not ok:
        raise HTTPException(status_code=409, detail="Username already exists")
    return {"status": "created", "username": req.username}


# ── Queue ─────────────────────────────────────────────────────────────────────

@router.get("/queue/next")
async def get_next(username: str = Depends(_require_auth)):
    item, prefetch_url = await db.get_next_unvalidated(username)
    stats = await db.get_validation_stats()
    return {"item": item, "stats": stats, "prefetch_url": prefetch_url}


@router.get("/queue/stats")
async def get_stats(username: str = Depends(_require_auth)):
    return await db.get_validation_stats()


# ── Screenshot ────────────────────────────────────────────────────────────────

async def _resolve_pdf_path(mun: str, zona: str, puesto: str, mesa: int, corp: str):
    conn = await db.get_db()
    rows = await conn.execute_fetchall(
        """SELECT filepath FROM e14_downloads
           WHERE municipio_cod=? AND zona_cod=? AND puesto_cod=?
             AND mesa=? AND corporacion=? LIMIT 1""",
        (mun, zona, puesto, mesa, corp.upper()),
    )
    if not rows:
        raise HTTPException(status_code=404, detail="PDF file not found")
    raw = rows[0]["filepath"]
    # Support both absolute paths (Railway /persist/...) and legacy relative paths
    from pathlib import Path as _Path
    import os as _os
    if _os.path.isabs(raw):
        full_path = _Path(raw).resolve()
    else:
        full_path = (BASE_DIR / raw).resolve()
    # Security: must be inside E14_DOWNLOADS_DIR or BASE_DIR
    allowed = [E14_DOWNLOADS_DIR.resolve(), BASE_DIR.resolve()]
    if not any(_safe_within(full_path, root) for root in allowed):
        raise HTTPException(status_code=400, detail="Invalid path")
    if not full_path.exists():
        raise HTTPException(status_code=404, detail="PDF file not found")
    return full_path


@router.get("/screenshot/{mun}/{zona}/{puesto}/{mesa}/{corp}")
async def get_screenshot(mun: str, zona: str, puesto: str, mesa: int, corp: str):
    import asyncio
    from backend.services.screenshot import render_pacto_crop

    # ── Cache check FIRST — zero DB queries on hit ─────────────────────────
    url_key = f"{mun}:{zona}:{puesto}:{mesa}:{corp.upper()}"
    png = _SCREENSHOT_CACHE.get(url_key)
    if png is not None:
        return Response(content=png, media_type="image/png",
                        headers={"Cache-Control": "public, max-age=3600"})

    # ── Cache miss: resolve path + render in dedicated thread pool ─────────
    full_path = await _resolve_pdf_path(mun, zona, puesto, mesa, corp)
    override = await db.get_crop_override(mun, zona, puesto, mesa, corp.upper())

    loop = asyncio.get_event_loop()
    png = await loop.run_in_executor(
        _SCREENSHOT_POOL,
        lambda: render_pacto_crop(str(full_path), corp, override=override)
    )
    if len(_SCREENSHOT_CACHE) >= _SCREENSHOT_CACHE_MAX:
        _SCREENSHOT_CACHE.pop(next(iter(_SCREENSHOT_CACHE)))
    _SCREENSHOT_CACHE[url_key] = png

    return Response(content=png, media_type="image/png",
                    headers={"Cache-Control": "public, max-age=3600"})


@router.get("/fullpage/{mun}/{zona}/{puesto}/{mesa}/{corp}")
async def get_fullpage(mun: str, zona: str, puesto: str, mesa: int, corp: str):
    """Return the full PDF page image for the crop editor."""
    from backend.services.screenshot import render_full_page
    full_path = await _resolve_pdf_path(mun, zona, puesto, mesa, corp)
    return Response(content=render_full_page(str(full_path), corp), media_type="image/png")


# ── Submit single validation ──────────────────────────────────────────────────

@router.post("/submit")
async def submit(req: SubmitRequest, username: str = Depends(_require_auth)):
    if req.action not in ("approved", "corrected"):
        raise HTTPException(status_code=400, detail="action must be 'approved' or 'corrected'")
    if req.action == "corrected" and req.corrected_ph_votes is None:
        raise HTTPException(status_code=400, detail="corrected_ph_votes required")

    try:
        await db.submit_validation({
            "municipio_cod": req.municipio_cod,
            "zona_cod": req.zona_cod,
            "puesto_cod": req.puesto_cod,
            "mesa": req.mesa,
            "corporacion": req.corporacion.upper(),
            "validated_by": username,
            "action": req.action,
            "corrected_ph_votes": req.corrected_ph_votes,
        })
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    await db.release_claim(username)

    # Evaluate discrepancy — fires only when both SEN and CAM are validated
    await alert_engine.evaluate_mesa(
        req.municipio_cod, req.zona_cod, req.puesto_cod, req.mesa
    )
    return {"status": "ok", "action": req.action}


# ── Novelty ───────────────────────────────────────────────────────────────────

@router.post("/novelty")
async def report_novelty(req: NoveltyRequest, username: str = Depends(_require_auth)):
    await db.add_novelty_note(
        req.municipio_cod, req.zona_cod, req.puesto_cod,
        req.mesa, req.corporacion.upper(), username, req.note,
    )
    await db.release_claim(username)
    await event_bus.publish("alert_created", {
        "municipio_cod": req.municipio_cod,
        "type": "novelty_report",
        "severity": "info",
    })
    return {"status": "ok"}


@router.get("/admin/validations")
async def list_validations(search: str = "", admin_token: str = ""):
    """List all validations for admin review (requires admin token)."""
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")
    return await db.get_all_validations(search)


class AdminCorrectRequest(BaseModel):
    validation_id: int
    corrected_ph_votes: int
    admin_token: str


@router.post("/admin/correct-validation")
async def admin_correct_validation(req: AdminCorrectRequest):
    """Admin override of a validator's submitted value."""
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(req.admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")
    ok = await db.admin_correct_validation(req.validation_id, req.corrected_ph_votes, "admin")
    if not ok:
        raise HTTPException(status_code=404, detail="Validación no encontrada")
    # Look up the mesa to re-evaluate discrepancy
    conn = await db.get_db()
    rows = await conn.execute_fetchall(
        "SELECT municipio_cod, zona_cod, puesto_cod, mesa FROM manual_validations WHERE id = ?",
        (req.validation_id,),
    )
    if rows:
        r = rows[0]
        await alert_engine.evaluate_mesa(r["municipio_cod"], r["zona_cod"], r["puesto_cod"], r["mesa"])
    return {"status": "ok", "validation_id": req.validation_id, "new_value": req.corrected_ph_votes}


@router.post("/undo")
async def undo_validation(username: str = Depends(_require_auth)):
    """Undo the validator's most recent submission and return that item."""
    item = await db.undo_last_validation(username)
    stats = await db.get_validation_stats()
    return {"item": item, "stats": stats}


@router.get("/progress")
async def get_progress():
    """Public progress dashboard: E14 downloaded / processed / validated."""
    return await db.get_e14_progress()


@router.get("/admin/users")
async def list_users(admin_token: str = ""):
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token invalido")
    return await db.list_users()


@router.post("/admin/users/deactivate")
async def deactivate_validation_users(req: AdminUsersRequest):
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(req.admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token invalido")
    result = await db.deactivate_users(req.usernames)
    return {"status": "ok", **result}


@router.delete("/admin/users")
async def delete_validation_users(req: AdminUsersRequest):
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(req.admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token invalido")
    result = await db.delete_users(req.usernames)
    return {"status": "ok", **result}


@router.delete("/admin/sessions")
async def clear_all_sessions(admin_token: str = ""):
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")
    conn = await db.get_db()
    await conn.execute("DELETE FROM sessions")
    await conn.execute("DELETE FROM queue_claims")
    await conn.commit()
    return {"status": "ok", "message": "Todas las sesiones y claims cerrados"}


@router.get("/novedades")
async def get_novedades():
    return await db.get_novelty_reports()


class ResolveNoveltyRequest(BaseModel):
    admin_token: str
    corrected_ph_votes: int | None = None


@router.post("/novedades/{novelty_id}/resolve")
async def resolve_novelty(novelty_id: int, req: ResolveNoveltyRequest):
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(req.admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")
    ok = await db.resolve_novelty(novelty_id, "admin", req.corrected_ph_votes)
    if not ok:
        raise HTTPException(status_code=404, detail="Novedad no encontrada")
    if req.corrected_ph_votes is not None:
        conn = await db.get_db()
        rows = await conn.execute_fetchall(
            "SELECT municipio_cod, zona_cod, puesto_cod, mesa FROM manual_validations WHERE id=?",
            (novelty_id,),
        )
        if rows:
            r = rows[0]
            await alert_engine.evaluate_mesa(r["municipio_cod"], r["zona_cod"], r["puesto_cod"], r["mesa"])
    return {"status": "ok"}


@router.post("/novedades/{novelty_id}/unresolve")
async def unresolve_novelty(novelty_id: int, req: ResolveNoveltyRequest):
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(req.admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")
    await db.unresolve_novelty(novelty_id)
    return {"status": "ok"}


class DeleteNoveltyRequest(BaseModel):
    admin_token: str


class PublicMunicipioExportRequest(BaseModel):
    municipio_cod: str


@router.delete("/novedades/{novelty_id}")
async def delete_novelty_and_reprocess(novelty_id: int, req: DeleteNoveltyRequest):
    """Admin: delete a novelty + its OCR result + download record + PDF so the poller re-downloads it."""
    from pathlib import Path as _Path

    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(req.admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")

    conn = await db.get_db()
    rows = await conn.execute_fetchall(
        """SELECT mv.id as mv_id, mv.municipio_cod, mv.zona_cod, mv.puesto_cod,
                  mv.mesa, mv.corporacion,
                  r.id as result_id, r.download_id, d.filepath
           FROM manual_validations mv
           JOIN e14_results r ON (
               r.municipio_cod = mv.municipio_cod AND r.zona_cod = mv.zona_cod AND
               r.puesto_cod = mv.puesto_cod AND r.mesa = mv.mesa AND r.corporacion = mv.corporacion
           )
           LEFT JOIN e14_downloads d ON d.id = r.download_id
           WHERE mv.id = ?""",
        (novelty_id,)
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Novedad no encontrada")

    row = rows[0]

    # 1. Release queue claims
    await conn.execute(
        "DELETE FROM queue_claims WHERE municipio_cod=? AND zona_cod=? AND puesto_cod=? AND mesa=? AND corporacion=?",
        (row["municipio_cod"], row["zona_cod"], row["puesto_cod"], row["mesa"], row["corporacion"])
    )
    # 2. Delete the novelty
    await conn.execute("DELETE FROM manual_validations WHERE id = ?", (row["mv_id"],))
    # 3. Delete OCR result
    await conn.execute("DELETE FROM e14_results WHERE id = ?", (row["result_id"],))
    # 4. Delete download record
    if row["download_id"]:
        await conn.execute("DELETE FROM e14_downloads WHERE id = ?", (row["download_id"],))
    await conn.commit()

    # 5. Remove PDF from disk
    deleted_file = False
    if row["filepath"]:
        try:
            _Path(row["filepath"]).unlink(missing_ok=True)
            deleted_file = True
        except Exception:
            pass

    return {
        "status": "ok",
        "deleted_file": deleted_file,
        "message": "Novedad eliminada — PDF será re-descargado y procesado en el próximo ciclo.",
    }


@router.get("/admin/debug-statuses")
async def debug_gql_statuses(admin_token: str = "", mun: str = "001", corp: str = "001", zone: str = "99"):
    """Admin: dump all GQL records for a mun+zone to compare standCode vs DB puesto_cod."""
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")

    from backend.services.remote_poller import RemotePoller, _QUERY_TRANSMISSION_CODES_BY_MUN
    import httpx

    poller = RemotePoller()
    all_nodes = []
    async with httpx.AsyncClient(follow_redirects=True) as client:
        if not await poller._get_creds(client):
            raise HTTPException(status_code=503, detail="No se pudieron obtener credenciales AWS")
        for status in (3, 11):
            data = await poller._gql(
                client, _QUERY_TRANSMISSION_CODES_BY_MUN,
                {"first": 5000, "status": status, "dept": "01", "corp": corp, "mun": mun},
            )
            if data:
                nodes = data.get("allTransmissionCodes", {}).get("nodes") or []
                all_nodes.extend(nodes)

    # Filter to requested zone
    filtered = [n for n in all_nodes if str(n.get("idZoneCode") or "") == zone]
    # Compare with DB
    conn = await db.get_db()
    db_rows = await conn.execute_fetchall(
        "SELECT puesto_cod, nombre FROM puestos WHERE municipio_cod=? AND zona_cod=?",
        (mun, zone)
    )
    db_puestos = {r["puesto_cod"]: r["nombre"] for r in db_rows}
    # Group GQL by standCode
    from collections import defaultdict
    by_stand: dict = defaultdict(list)
    for n in filtered:
        by_stand[str(n.get("standCode") or "")].append({
            "mesa": n.get("numberStand"),
            "status": n.get("idTransmissionCodeStatus"),
            "file": n.get("expectedName"),
        })

    return {
        "gql_standCodes": dict(by_stand),
        "db_puesto_cods": db_puestos,
        "gql_total_zone": len(filtered),
    }


@router.get("/admin/missing-mesas")
async def get_missing_mesas(admin_token: str = "", limit: int = 10):
    """Admin: mesas in catalog with no e14 download yet."""
    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")
    conn = await db.get_db()
    rows = await conn.execute_fetchall(f"""
        SELECT p.municipio_cod, p.zona_cod, p.puesto_cod, p.municipio, p.nombre,
               p.mesas,
               COUNT(d.id) as descargas,
               (p.mesas * 2 - COUNT(d.id)) as faltantes
        FROM puestos p
        LEFT JOIN e14_downloads d ON d.municipio_cod = p.municipio_cod
            AND d.zona_cod = p.zona_cod AND d.puesto_cod = p.puesto_cod
        WHERE p.departamento = 'ANTIOQUIA'
        GROUP BY p.municipio_cod, p.zona_cod, p.puesto_cod
        HAVING faltantes > 0
        ORDER BY faltantes DESC
        LIMIT {limit}
    """)
    return [dict(r) for r in rows]


@router.get("/admin/alertas/export-excel")
async def export_alertas_excel(admin_token: str = ""):
    """Admin: export all unresolved vote_discrepancy alerts as Excel (compatible with generate_reclamaciones scripts)."""
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime as _dt

    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")

    conn = await db.get_db()
    rows = await conn.execute_fetchall(
        """
        SELECT
            a.municipio_cod, a.zona_cod, a.puesto_cod, a.mesa,
            a.discrepancy_pct,
            p.municipio, p.nombre as puesto_nombre, p.departamento,
            -- Use corrected votes if available, else OCR
            COALESCE(mv_sen.corrected_ph_votes, r_sen.ph_total_votos) as votos_sen,
            COALESCE(mv_cam.corrected_ph_votes, r_cam.ph_total_votos) as votos_cam,
            r_sen.ph_total_votos as sen_ocr, r_cam.ph_total_votos as cam_ocr,
            r_sen.votos_urna as sen_votos_urna, r_cam.votos_urna as cam_votos_urna,
            r_sen.ocr_confidence as sen_conf, r_cam.ocr_confidence as cam_conf,
            mv_sen.corrected_ph_votes as sen_corrected, mv_cam.corrected_ph_votes as cam_corrected,
            mv_sen.action as sen_action, mv_cam.action as cam_action
        FROM alerts a
        LEFT JOIN puestos p ON p.municipio_cod = a.municipio_cod
            AND p.zona_cod = a.zona_cod AND p.puesto_cod = a.puesto_cod
        LEFT JOIN e14_results r_sen ON r_sen.municipio_cod = a.municipio_cod
            AND r_sen.zona_cod = a.zona_cod AND r_sen.puesto_cod = a.puesto_cod
            AND r_sen.mesa = a.mesa AND r_sen.corporacion = 'SEN'
        LEFT JOIN e14_results r_cam ON r_cam.municipio_cod = a.municipio_cod
            AND r_cam.zona_cod = a.zona_cod AND r_cam.puesto_cod = a.puesto_cod
            AND r_cam.mesa = a.mesa AND r_cam.corporacion = 'CAM'
        LEFT JOIN manual_validations mv_sen ON mv_sen.municipio_cod = a.municipio_cod
            AND mv_sen.zona_cod = a.zona_cod AND mv_sen.puesto_cod = a.puesto_cod
            AND mv_sen.mesa = a.mesa AND mv_sen.corporacion = 'SEN'
        LEFT JOIN manual_validations mv_cam ON mv_cam.municipio_cod = a.municipio_cod
            AND mv_cam.zona_cod = a.zona_cod AND mv_cam.puesto_cod = a.puesto_cod
            AND mv_cam.mesa = a.mesa AND mv_cam.corporacion = 'CAM'
        WHERE a.is_resolved = 0 AND a.alert_type = 'vote_discrepancy'
        ORDER BY p.municipio, a.mesa
        """
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alertas"

    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ctr = Alignment(horizontal="center", vertical="center")

    headers = [
        "Llave DMZPM",
        "nombre_municipio", "codigo_zona", "nombre_puesto", "Mesa",
        "Votos Camara", "Votos Senado", "Dif Cam-Sen",
        "Departamento", "municipio_cod", "zona_cod", "puesto_cod",
        "SEN OCR orig", "CAM OCR orig",
        "SEN corregido", "CAM corregido",
        "SEN accion", "CAM accion",
        "SEN votos urna", "CAM votos urna",
        "SEN confianza OCR", "CAM confianza OCR",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = ctr

    for row in rows:
        mun = (row["municipio_cod"] or "").zfill(3)
        zona = (row["zona_cod"] or "").zfill(2)
        puesto = (row["puesto_cod"] or "").zfill(2)
        mesa = row["mesa"] or 0
        # Reconstruct Llave DMZPM: 1 + mun(3) + zona(2) + puesto(2) + mesa(3)
        llave = f"1{mun}{zona}{puesto}{str(mesa).zfill(3)}"
        votos_sen = row["votos_sen"]
        votos_cam = row["votos_cam"]
        dif = (votos_cam or 0) - (votos_sen or 0) if (votos_sen is not None and votos_cam is not None) else None
        ws.append([
            llave,
            row["municipio"] or row["municipio_cod"],
            zona,
            row["puesto_nombre"] or "",
            mesa,
            votos_cam, votos_sen, dif,
            row["departamento"] or "",
            row["municipio_cod"], row["zona_cod"], row["puesto_cod"],
            row["sen_ocr"], row["cam_ocr"],
            row["sen_corrected"], row["cam_corrected"],
            row["sen_action"], row["cam_action"],
            row["sen_votos_urna"], row["cam_votos_urna"],
            row["sen_conf"], row["cam_conf"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"alertas_discrepancia_{_dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/admin/alertas/generate-zip")
async def generate_alertas_zip(req: DeleteNoveltyRequest):
    """Admin: build ZIP of all vote_discrepancy PDFs and save to /persist/exports/. Returns filename."""
    import asyncio as _asyncio, io, zipfile
    from pathlib import Path as _Path
    from datetime import datetime as _dt

    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(req.admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")

    conn = await db.get_db()
    rows = await conn.execute_fetchall(
        """
        SELECT a.municipio_cod, a.zona_cod, a.puesto_cod, a.mesa,
               a.discrepancy_pct,
               d_sen.filepath as sen_path, d_sen.filename as sen_filename,
               d_cam.filepath as cam_path, d_cam.filename as cam_filename
        FROM alerts a
        LEFT JOIN e14_downloads d_sen ON d_sen.municipio_cod = a.municipio_cod
            AND d_sen.zona_cod = a.zona_cod AND d_sen.puesto_cod = a.puesto_cod
            AND d_sen.mesa = a.mesa AND d_sen.corporacion = 'SEN'
        LEFT JOIN e14_downloads d_cam ON d_cam.municipio_cod = a.municipio_cod
            AND d_cam.zona_cod = a.zona_cod AND d_cam.puesto_cod = a.puesto_cod
            AND d_cam.mesa = a.mesa AND d_cam.corporacion = 'CAM'
        WHERE a.is_resolved = 0 AND a.alert_type = 'vote_discrepancy'
        ORDER BY a.discrepancy_pct DESC
        """
    )

    rows_list = list(rows)
    exports_dir = _EXPORTS_DIR
    exports_dir.mkdir(parents=True, exist_ok=True)
    fname = f"alertas_discrepancia_{_dt.now().strftime('%Y%m%d_%H%M%S')}.zip"
    out_path = exports_dir / fname

    def _build():
        added: set[str] = set()
        with zipfile.ZipFile(str(out_path), "w", zipfile.ZIP_DEFLATED) as zf:
            for row in rows_list:
                for path_field, name_field in [("sen_path", "sen_filename"), ("cam_path", "cam_filename")]:
                    filepath = row[path_field]
                    filename = row[name_field]
                    if not filepath or filepath in added:
                        continue
                    p = _Path(filepath)
                    if p.exists():
                        mun = row["municipio_cod"]
                        mesa = row["mesa"]
                        pct = row["discrepancy_pct"]
                        arcname = f"{pct:.0f}pct/{mun}_mesa{mesa:03d}_{filename or p.name}"
                        zf.write(str(p), arcname)
                        added.add(filepath)
        return str(out_path), len(added)

    loop = _asyncio.get_event_loop()
    zip_path, count = await loop.run_in_executor(None, _build)
    return {"status": "ok", "filename": fname, "files": count, "path": zip_path}


@router.get("/admin/alertas/download-zip")
async def download_alertas_zip(admin_token: str = "", filename: str = ""):
    """Admin: download a previously generated ZIP by filename."""
    from pathlib import Path as _Path

    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")
    if not filename or "/" in filename or "\\" in filename or not filename.endswith(".zip"):
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")

    exports_dir = _EXPORTS_DIR
    zip_path = exports_dir / filename
    if not zip_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado — usa /generate-zip primero")

    def _iter_file():
        with open(zip_path, "rb") as f:
            while chunk := f.read(65536):
                yield chunk

    return StreamingResponse(
        _iter_file(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/public-exports/share/{share_token}/municipios")
async def list_public_export_municipios(share_token: str):
    _require_public_export_access(share_token)

    conn = await db.get_db()
    rows = await conn.execute_fetchall(
        """
        SELECT d.municipio_cod,
               COALESCE(MAX(NULLIF(TRIM(p.municipio), '')), d.municipio_cod) as municipio,
               COUNT(*) as pdf_count
        FROM e14_downloads d
        LEFT JOIN puestos p
          ON p.municipio_cod = d.municipio_cod
         AND p.zona_cod = d.zona_cod
         AND p.puesto_cod = d.puesto_cod
        GROUP BY d.municipio_cod
        ORDER BY municipio, d.municipio_cod
        """
    )
    return [
        {
            "municipio_cod": row["municipio_cod"],
            "municipio": row["municipio"],
            "pdf_count": row["pdf_count"],
        }
        for row in rows
    ]


@router.get("/public-exports/share/{share_token}/{filename}")
async def download_public_export(share_token: str, filename: str):
    """Download a share-only ZIP from /persist/exports using the share token."""
    _require_public_export_access(share_token)
    if (
        not filename
        or "/" in filename
        or "\\" in filename
        or not filename.endswith(".zip")
        or not filename.startswith("public_")
    ):
        raise HTTPException(status_code=400, detail="Nombre de archivo inválido")

    zip_path = _EXPORTS_DIR / filename
    if not zip_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    return _zip_stream_response(zip_path, filename)


@router.post("/public-exports/share/{share_token}/generate-municipio")
async def generate_public_municipio_export(share_token: str, req: PublicMunicipioExportRequest):
    _require_public_export_access(share_token)
    import asyncio as _asyncio
    import csv as _csv
    import os as _os
    import zipfile as _zipfile

    municipio_cod = (req.municipio_cod or "").strip()
    if not municipio_cod:
        raise HTTPException(status_code=400, detail="municipio_cod es requerido")

    conn = await db.get_db()
    rows = await conn.execute_fetchall(
        """
        SELECT d.municipio_cod, d.zona_cod, d.puesto_cod, d.mesa, d.corporacion,
               d.filename, d.filepath,
               p.municipio, p.nombre as puesto_nombre
        FROM e14_downloads d
        LEFT JOIN puestos p
          ON p.municipio_cod = d.municipio_cod
         AND p.zona_cod = d.zona_cod
         AND p.puesto_cod = d.puesto_cod
        WHERE d.municipio_cod = ?
        ORDER BY d.zona_cod, d.puesto_cod, d.mesa, d.corporacion
        """,
        (municipio_cod,),
    )
    if not rows:
        raise HTTPException(status_code=404, detail="No hay PDFs descargados para ese municipio")

    rows_list = [dict(row) for row in rows]
    municipio_nombre = next(
        (str(row.get("municipio") or "").strip() for row in rows_list if row.get("municipio")),
        municipio_cod,
    )
    municipio_slug = _clean_zip_segment(municipio_nombre, f"municipio_{municipio_cod}").lower().replace(" ", "_")
    root_dir = f"E14_{municipio_cod}-{_clean_zip_segment(municipio_nombre, municipio_cod)}"
    filename = f"public_e14_municipio_{municipio_cod}_{municipio_slug}.zip"
    zip_path = _EXPORTS_DIR / filename
    tmp_path = zip_path.with_suffix(".zip.tmp")

    def _build_zip() -> tuple[int, int, int]:
        _EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        if tmp_path.exists():
            tmp_path.unlink()

        files_added = 0
        missing_files = 0
        seen_paths: set[str] = set()
        manifest_buffer = io.StringIO()
        manifest_writer = _csv.writer(manifest_buffer)
        manifest_writer.writerow([
            "municipio_cod",
            "municipio",
            "zona_cod",
            "puesto_cod",
            "puesto_nombre",
            "mesa",
            "corporacion",
            "source_filename",
            "zip_path",
        ])

        with _zipfile.ZipFile(tmp_path, "w", compression=_zipfile.ZIP_STORED) as zf:
            for row in rows_list:
                resolved_path = _resolve_download_filepath(row.get("filepath"))
                if resolved_path is None or not resolved_path.exists():
                    missing_files += 1
                    continue

                source_key = str(resolved_path)
                if source_key in seen_paths:
                    continue
                seen_paths.add(source_key)

                zona_cod = str(row.get("zona_cod") or "").zfill(2)
                puesto_cod = str(row.get("puesto_cod") or "").zfill(2)
                mesa = int(row.get("mesa") or 0)
                corporacion = str(row.get("corporacion") or "").upper() or "PDF"
                puesto_nombre = _clean_zip_segment(row.get("puesto_nombre"), f"PUESTO_{puesto_cod}")
                zip_member = (
                    f"{root_dir}/"
                    f"{zona_cod}-ZONA/"
                    f"{puesto_cod}-{puesto_nombre}/"
                    f"MESA_{mesa:03d}_{corporacion}.pdf"
                )
                zf.write(str(resolved_path), zip_member)
                manifest_writer.writerow([
                    municipio_cod,
                    municipio_nombre,
                    zona_cod,
                    puesto_cod,
                    row.get("puesto_nombre") or "",
                    mesa,
                    corporacion,
                    row.get("filename") or resolved_path.name,
                    zip_member,
                ])
                files_added += 1

            zf.writestr("MANIFEST.csv", manifest_buffer.getvalue())

        _os.replace(tmp_path, zip_path)
        return files_added, missing_files, zip_path.stat().st_size

    loop = _asyncio.get_event_loop()
    files_added, missing_files, size_bytes = await loop.run_in_executor(None, _build_zip)

    if files_added == 0:
        if zip_path.exists():
            zip_path.unlink()
        raise HTTPException(status_code=404, detail="No hay archivos disponibles para exportar en ese municipio")

    return {
        "status": "ok",
        "municipio_cod": municipio_cod,
        "municipio": municipio_nombre,
        "filename": filename,
        "public_url": f"/api/validar/public-exports/share/{share_token}/{filename}",
        "files": files_added,
        "missing_files": missing_files,
        "size_bytes": size_bytes,
    }


@router.post("/admin/novedades/purge-bad-scan")
async def purge_bad_scan_novelties(req: DeleteNoveltyRequest):
    """Admin: delete all 'Mal escaneado' novelties + their OCR results + PDFs so they get re-downloaded."""
    from pathlib import Path as _Path

    if not VALIDATE_SETUP_TOKEN or not secrets.compare_digest(req.admin_token, VALIDATE_SETUP_TOKEN):
        raise HTTPException(status_code=403, detail="Token inválido")

    conn = await db.get_db()
    rows = await conn.execute_fetchall(
        """SELECT mv.id as mv_id, mv.municipio_cod, mv.zona_cod, mv.puesto_cod,
                  mv.mesa, mv.corporacion,
                  r.id as result_id, r.download_id, d.filepath
           FROM manual_validations mv
           LEFT JOIN e14_results r ON (
               r.municipio_cod = mv.municipio_cod AND r.zona_cod = mv.zona_cod AND
               r.puesto_cod = mv.puesto_cod AND r.mesa = mv.mesa AND r.corporacion = mv.corporacion
           )
           LEFT JOIN e14_downloads d ON d.id = r.download_id
           WHERE mv.novelty_note LIKE '%mal escaneado%'""",
    )
    if not rows:
        return {"status": "ok", "deleted": 0, "message": "No hay novedades de 'Mal escaneado'."}

    deleted_files: list[str] = []
    for row in rows:
        await conn.execute(
            "DELETE FROM queue_claims WHERE municipio_cod=? AND zona_cod=? AND puesto_cod=? AND mesa=? AND corporacion=?",
            (row["municipio_cod"], row["zona_cod"], row["puesto_cod"], row["mesa"], row["corporacion"])
        )
        await conn.execute("DELETE FROM manual_validations WHERE id = ?", (row["mv_id"],))
        if row["result_id"]:
            await conn.execute("DELETE FROM e14_results WHERE id = ?", (row["result_id"],))
        if row["download_id"]:
            await conn.execute("DELETE FROM e14_downloads WHERE id = ?", (row["download_id"],))
        if row["filepath"]:
            try:
                _Path(row["filepath"]).unlink(missing_ok=True)
                deleted_files.append(row["filepath"])
            except Exception:
                pass

    await conn.commit()

    return {
        "status": "ok",
        "deleted": len(rows),
        "deleted_files": len(deleted_files),
        "message": f"{len(rows)} novedades eliminadas — PDFs serán re-descargados en el próximo ciclo.",
    }


@router.get("/novedades/export")
async def export_novedades():
    """Download all novelty reports as Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = await db.get_novelty_reports()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Novedades"

    headers = [
        "ID", "Departamento", "Municipio", "Cód Municipio",
        "Zona", "Puesto", "Cód Puesto", "Mesa", "Corporación",
        "Validador", "Acción", "Fecha Validación",
        "Votos IA", "Votos Corregidos", "Votos Urna",
        "Nota de Novedad",
    ]

    # Header row styling
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, item in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=item.get("id"))
        ws.cell(row=row_idx, column=2, value=item.get("departamento") or "")
        ws.cell(row=row_idx, column=3, value=item.get("municipio") or item.get("municipio_cod"))
        ws.cell(row=row_idx, column=4, value=item.get("municipio_cod"))
        ws.cell(row=row_idx, column=5, value=item.get("zona_cod"))
        ws.cell(row=row_idx, column=6, value=item.get("puesto_nombre") or "")
        ws.cell(row=row_idx, column=7, value=item.get("puesto_cod"))
        ws.cell(row=row_idx, column=8, value=item.get("mesa"))
        ws.cell(row=row_idx, column=9, value=item.get("corporacion"))
        ws.cell(row=row_idx, column=10, value=item.get("validated_by"))
        ws.cell(row=row_idx, column=11, value=item.get("action"))
        ws.cell(row=row_idx, column=12, value=item.get("validated_at", "")[:19])
        ws.cell(row=row_idx, column=13, value=item.get("ai_ph_votes"))
        ws.cell(row=row_idx, column=14, value=item.get("corrected_ph_votes"))
        ws.cell(row=row_idx, column=15, value=item.get("votos_urna"))
        ws.cell(row=row_idx, column=16, value=item.get("novelty_note"))

    # Column widths
    widths = [6, 14, 20, 14, 8, 30, 10, 8, 12, 14, 10, 22, 10, 16, 11, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"novedades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
